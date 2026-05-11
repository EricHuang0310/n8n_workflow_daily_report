"""
session_view.py - 以 sessionID 為單位的每日對話明細表
========================================================

輸出格式: Excel (.xlsx)
  Sheet 1 「對話明細」:
    一列一個 sessionID, 欄位包含彙整後的對話 metadata 與內容。

使用方式:
  python session_view.py <execution_log.json> [--output session_report.xlsx]
"""

import json
import os
import sys
from collections import OrderedDict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils import (
    get_node_output,
    get_run_data,
    node_was_executed,
)


# ── 節點分類 ────────────────────────────────────────────

# 使用者送出訊息的入口節點（Webhook / Wait）
USER_MESSAGE_NODES = [
    "receive_message_API",
    "receive_confirmation",
    "receive_confirmation_again",
    "receive_confirmation_message",
    "receive_other_question",
    "receive_check_phone",
    "receive_speck_phone",
    "receive_confirm_phone",
]

# 機器人回覆訊息的節點（每個 respondToWebhook = 1 條 AI 訊息）
AI_RESPONSE_NODES = [
    "return_confirm",
    "return_negative",
    "return_other",
    "return_SMS",
    "return_IVR",
    "return_live_support1",
    "return_live_support2",
    "return_error_to_human",
    "retrun_negative_to_human",
    "retrun_other_to_human",
    "retrun_to_human",
    "retrun_to_human1",
    "retrun_to_human2",
    "return_check_phone",
    "return_confirm_phone",
    "return_other_question",
    "return_end",
    "return_speck_phone",
    "return_wait_intent",
]

# 有回答的終點節點
ANSWER_NODES = [
    "SMS_response",
    "IVR_response",
    "live_support_response1",
    "live_support_response2",
]

# 無答案/轉接 fallback 節點
NO_ANSWER_NODES = [
    "negative_response",
    "negative_response_to_human",
    "other_response",
    "other_response_to_human",
    "response_to_human",
    "response_to_human1",
    "response_to_human2",
    "error_to_human",
    "end_call",
]

# 走到 automation_router 各分支對應的「答案類型」(中文)
ACTION_LABELS = {
    "send_message": "發訊息(SMS)",
    "transfer_direct": "不核身轉接專員",
    "transfer_auth": "核身後轉接專員",
    "transfer_ivr_auth": "核身後轉接語音功能",
}


# ── 基礎輔助 ────────────────────────────────────────────

def _safe_get_body(run_data: dict) -> dict:
    """從 receive_message_API / clear_history_API 取出 body。"""
    for nm in ("receive_message_API", "clear_history_API"):
        out = get_node_output(run_data, nm)
        if out and isinstance(out, dict):
            body = out.get("body")
            if isinstance(body, dict):
                return body
    return {}


def _count_runs(run_data: dict, node_name: str) -> int:
    """節點被執行的次數（同一 execution 內可能 >1）。"""
    return len(run_data.get(node_name, []))


def _count_successful_runs(run_data: dict, node_name: str) -> int:
    """節點被執行成功的次數。Wait 節點 status=waiting 不計入。"""
    return sum(
        1 for ent in run_data.get(node_name, [])
        if ent.get("executionStatus") == "success"
    )


def _sum_execution_time_ms(run_data: dict) -> int:
    """加總本 execution 所有節點 executionTime（毫秒）。"""
    total = 0
    for entries in run_data.values():
        for ent in entries:
            v = ent.get("executionTime")
            if isinstance(v, (int, float)):
                total += int(v)
    return total


def _iter_node_outputs(run_data: dict, node_name: str):
    """yield 每一個 entry 的 json 輸出。"""
    for ent in run_data.get(node_name, []):
        try:
            main = ent.get("data", {}).get("main", [[]])
            if main and main[0]:
                yield main[0][0].get("json", {})
        except Exception:
            continue


def _collect_audio_files(run_data: dict) -> list[str]:
    """掃描所有 set 節點輸出，蒐集 audioFile 欄位的內容。"""
    audio = []
    for node_name, entries in run_data.items():
        for ent in entries:
            try:
                main = ent.get("data", {}).get("main", [[]])
                if not main or not main[0]:
                    continue
                j = main[0][0].get("json", {})
            except Exception:
                continue
            if not isinstance(j, dict):
                continue
            af = j.get("audioFile")
            if af is None:
                continue
            if isinstance(af, list):
                for f in af:
                    if f:
                        audio.append(str(f))
            elif isinstance(af, str):
                # 有時候會是 JSON 字串
                try:
                    parsed = json.loads(af)
                    if isinstance(parsed, list):
                        audio.extend(str(x) for x in parsed)
                    else:
                        audio.append(af)
                except Exception:
                    audio.append(af)
    return audio


# ── 單一 execution 的特徵抽取 ───────────────────────────

def _extract_execution_features(execution: dict) -> dict:
    """從單一 execution 抽出後續彙總需要的資訊。"""
    rd = get_run_data(execution)
    body = _safe_get_body(rd)

    # 使用者訊息 / AI 訊息計數 (僅計入 status=success)
    user_msg_count = sum(_count_successful_runs(rd, n) for n in USER_MESSAGE_NODES)
    ai_msg_count = sum(_count_successful_runs(rd, n) for n in AI_RESPONSE_NODES)

    # 有/無答案 (執行到即計入,不需 status=success)
    answered = sum(_count_runs(rd, n) for n in ANSWER_NODES)
    no_answer = sum(_count_runs(rd, n) for n in NO_ANSWER_NODES)

    # 使用者文字（按出現順序）
    user_texts = []
    if body.get("text"):
        user_texts.append(str(body["text"]))
    for nm in (
        "receive_confirmation",
        "receive_confirmation_again",
        "receive_confirmation_message",
        "receive_other_question",
        "receive_check_phone",
        "receive_speck_phone",
        "receive_confirm_phone",
    ):
        for out in _iter_node_outputs(rd, nm):
            b = out.get("body") if isinstance(out, dict) else None
            if isinstance(b, dict) and b.get("text"):
                user_texts.append(str(b["text"]))

    # 意圖（標準問題）
    intents = []
    response_models = []
    for out in _iter_node_outputs(rd, "intent_identification"):
        for item in out.get("output", []) or []:
            it = item.get("intent")
            if it:
                intents.append(it)
            rm = item.get("responseModel")
            if rm and rm not in response_models:
                response_models.append(rm)

    # 觸發條件 / 標準答案 / 答案類型
    subsequent_actions = []
    standard_answers = []
    for out in _iter_node_outputs(rd, "positive_metadata"):
        meta = out.get("metadata", {}) if isinstance(out, dict) else {}
        sa = meta.get("subsequentActions")
        if sa:
            subsequent_actions.append(str(sa))
        ans = meta.get("standardAnswer")
        if ans:
            standard_answers.append(str(ans))

    # 答案類型: 看走到哪個終點
    answer_type = None
    for nm, label in (
        ("SMS_response", "發訊息(SMS)"),
        ("IVR_response", "核身後轉接語音功能"),
        ("live_support_response1", "不核身轉接專員"),
        ("live_support_response2", "核身後轉接專員"),
    ):
        if node_was_executed(rd, nm):
            answer_type = label
            break
    if answer_type is None:
        # 後備: 從 subsequentActions 推
        if subsequent_actions:
            answer_type = subsequent_actions[-1]

    # 音檔
    audio_files = _collect_audio_files(rd)

    # 節點時間
    sum_node_ms = _sum_execution_time_ms(rd)

    # 起訖節點: 取 runData 內最早 / 最晚 startTime 對應的節點
    first_node = last_node = None
    first_t = last_t = None
    for nm, entries in rd.items():
        for ent in entries:
            st = ent.get("startTime")
            if not isinstance(st, (int, float)):
                continue
            if first_t is None or st < first_t:
                first_t, first_node = st, nm
            if last_t is None or st > last_t:
                last_t, last_node = st, nm

    # 用 result 內最終節點覆寫 last_node
    last_node_from_meta = (
        execution.get("data", {}).get("resultData", {}).get("lastNodeExecuted")
    )
    if last_node_from_meta:
        last_node = last_node_from_meta

    return {
        "execution_id": execution.get("id"),
        "status": execution.get("status"),
        "started_at": execution.get("startedAt"),
        "stopped_at": execution.get("stoppedAt"),
        "session_id": body.get("sessionID"),
        "customer_id": body.get("customerID"),
        "user_msg_count": user_msg_count,
        "ai_msg_count": ai_msg_count,
        "answered_count": answered,
        "no_answer_count": no_answer,
        "user_texts": user_texts,
        "intents": intents,
        "response_models": response_models,
        "subsequent_actions": subsequent_actions,
        "standard_answers": standard_answers,
        "answer_type": answer_type,
        "audio_files": audio_files,
        "sum_node_ms": sum_node_ms,
        "first_node": first_node,
        "last_node": last_node,
    }


# ── 跨 execution 彙總 (同 sessionID) ─────────────────────

def _merge_by_session(features: list[dict]) -> list[dict]:
    """同 sessionID 的 execution 合併成一列。"""
    grouped: "OrderedDict[str, list[dict]]" = OrderedDict()
    for f in features:
        sid = f.get("session_id") or "(unknown)"
        grouped.setdefault(sid, []).append(f)

    rows = []
    for sid, items in grouped.items():
        # 依 started_at 排序
        items_sorted = sorted(items, key=lambda x: x.get("started_at") or "")

        customer_id = next(
            (x.get("customer_id") for x in items_sorted if x.get("customer_id")), None
        )
        started_list = [x["started_at"] for x in items_sorted if x.get("started_at")]
        stopped_list = [x["stopped_at"] for x in items_sorted if x.get("stopped_at")]

        # 累加計數
        user_msg = sum(x.get("user_msg_count", 0) for x in items_sorted)
        ai_msg = sum(x.get("ai_msg_count", 0) for x in items_sorted)
        answered = sum(x.get("answered_count", 0) for x in items_sorted)
        no_ans = sum(x.get("no_answer_count", 0) for x in items_sorted)

        # 串接文字（保序、去除空）
        def _join_unique(values, sep=" | "):
            seen, out = set(), []
            for v in values:
                if v is None:
                    continue
                s = str(v).strip()
                if not s or s in seen:
                    continue
                seen.add(s)
                out.append(s)
            return sep.join(out)

        user_texts = []
        intents = []
        resp_models = []
        subsequent = []
        standard_answers = []
        answer_types = []
        audio_files = []
        sum_node_ms = 0
        statuses = []
        exec_ids = []

        for x in items_sorted:
            user_texts.extend(x.get("user_texts") or [])
            intents.extend(x.get("intents") or [])
            resp_models.extend(x.get("response_models") or [])
            subsequent.extend(x.get("subsequent_actions") or [])
            standard_answers.extend(x.get("standard_answers") or [])
            if x.get("answer_type"):
                answer_types.append(x["answer_type"])
            audio_files.extend(x.get("audio_files") or [])
            sum_node_ms += x.get("sum_node_ms", 0)
            if x.get("status"):
                statuses.append(x["status"])
            if x.get("execution_id") is not None:
                exec_ids.append(str(x["execution_id"]))

        first_node = items_sorted[0].get("first_node")
        last_node = items_sorted[-1].get("last_node")

        rows.append({
            "會話編號": sid,
            "customerID": customer_id,
            "進入時間": min(started_list) if started_list else None,
            "最後記錄時間": max(stopped_list) if stopped_list else None,
            "訊息量(用戶)": user_msg,
            "訊息量(AI)": ai_msg,
            "總訊息數": user_msg + ai_msg,
            "有回答(次)": answered,
            "無答案(次)": no_ans,
            "TxnID": ",".join(exec_ids),
            "測試者提問": _join_unique(user_texts),
            "標準問題": _join_unique(intents),
            "答案類型": _join_unique(answer_types),
            "機器人回答來源": _join_unique(resp_models),
            "台新腦回答": _join_unique(standard_answers),
            "情境狀態": statuses[-1] if statuses else None,
            "情境節點(從哪裡來)": first_node,
            "情境節點(往哪裡去)": last_node,
            "觸發條件": _join_unique(subsequent),
            "語音檔列表": ",".join(audio_files),
            "各節點停留時間總和(秒)": round(sum_node_ms / 1000.0, 3),
        })

    return rows


# ── Excel 輸出 ──────────────────────────────────────────

COLUMNS = [
    "會話編號",
    "customerID",
    "進入時間",
    "最後記錄時間",
    "訊息量(用戶)",
    "訊息量(AI)",
    "總訊息數",
    "有回答(次)",
    "無答案(次)",
    "TxnID",
    "測試者提問",
    "標準問題",
    "答案類型",
    "機器人回答來源",
    "台新腦回答",
    "情境狀態",
    "情境節點(從哪裡來)",
    "情境節點(往哪裡去)",
    "觸發條件",
    "語音檔列表",
    "各節點停留時間總和(秒)",
]

# 各欄寬度（字元）
COLUMN_WIDTHS = {
    "會話編號": 22,
    "customerID": 18,
    "進入時間": 22,
    "最後記錄時間": 22,
    "訊息量(用戶)": 12,
    "訊息量(AI)": 12,
    "總訊息數": 10,
    "有回答(次)": 12,
    "無答案(次)": 12,
    "TxnID": 18,
    "測試者提問": 40,
    "標準問題": 24,
    "答案類型": 18,
    "機器人回答來源": 18,
    "台新腦回答": 50,
    "情境狀態": 12,
    "情境節點(從哪裡來)": 22,
    "情境節點(往哪裡去)": 22,
    "觸發條件": 22,
    "語音檔列表": 50,
    "各節點停留時間總和(秒)": 18,
}


def _write_xlsx(rows: list[dict], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "對話明細"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    # Header
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Body
    for r, row in enumerate(rows, 2):
        for c, col in enumerate(COLUMNS, 1):
            v = row.get(col)
            ws.cell(row=r, column=c, value=v).alignment = body_align

    # 欄寬
    for c, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(c)].width = COLUMN_WIDTHS.get(col, 18)

    # 凍結首列 + 首欄, 啟用自動篩選
    ws.freeze_panes = "B2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"

    wb.save(output_path)


# ── 主流程 ──────────────────────────────────────────────

def build_session_report(log_filepath: str, output_path: str) -> int:
    with open(log_filepath, "r", encoding="utf-8") as f:
        raw = json.load(f)
    if isinstance(raw, dict):
        executions = raw.get("data", [raw])
    elif isinstance(raw, list):
        executions = raw
    else:
        executions = [raw]

    print(f"[SessionView] 讀入 {len(executions)} 筆 execution")

    features = [_extract_execution_features(ex) for ex in executions]
    rows = _merge_by_session(features)

    print(f"[SessionView] 合併後 {len(rows)} 筆 session")
    _write_xlsx(rows, output_path)
    print(f"[SessionView] 已輸出: {output_path}")
    return len(rows)


def main():
    args = sys.argv[1:]
    log_file = None
    output_file = None

    i = 0
    while i < len(args):
        if args[i] == "--output" and i + 1 < len(args):
            output_file = args[i + 1]
            i += 2
        elif not args[i].startswith("-"):
            log_file = args[i]
            i += 1
        else:
            i += 1

    if not log_file:
        print("使用方式: python session_view.py <execution_log.json> [--output session_report.xlsx]")
        sys.exit(1)

    if not os.path.exists(log_file):
        print(f"錯誤: 找不到檔案 {log_file}")
        sys.exit(1)

    if not output_file:
        today = datetime.now().strftime("%Y%m%d")
        output_file = f"SessionDailyReport_{today}.xlsx"

    build_session_report(log_file, output_file)


if __name__ == "__main__":
    main()

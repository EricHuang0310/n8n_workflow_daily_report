"""
excel_writer.py - 將報表 dict 輸出為 Excel (.xlsx) 檔案
==========================================================
入口函式:
    build_excel_report(report: dict, output_path: str) -> None

輸出結構:
    Sheet 1: 總覽
        - 報表 metadata (產出時間、資料範圍、execution 統計)
        - 逐模組呈現 PDF 要求的彙總指標 (key-value + 分布子表)
    Sheet 2..N: 每個報表模組獨立一張 Sheet
        - 純明細表格: 一列一筆 execution
        - 欄位為所有 record 的 key 聯集
        - node_18 / node_19 額外在頂端標註「workflow 尚未完整實作」

依賴: openpyxl
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 欄位中文標籤對照表 ────────────────────────────────────────────
# 用於總覽 Sheet 呈現彙總指標時，將英文 key 轉為 PDF 上的中文敘述
FIELD_LABELS = {
    # 共用
    "total_count": "進入此節點之數量",
    "total_occurrences": "發生次數",
    "total_error_count": "ERROR 發生次數",
    "total_timeout_count": "Time Out 發生次數",
    "avg_stay_duration_sec": "平均客戶停留時間(秒)",
    "avg_intent_identification_time_ms": "平均意圖辨識時間(毫秒)",
    "unique_customer_count": "獨立客戶數",
    # node_08
    "clear_question_count": "明確問題的筆數",
    "ambiguous_question_count": "模糊問題的筆數",
    # node_09
    "positive_count": "客戶正向回覆的筆數",
    "negative_count": "客戶負向回覆的筆數",
    "other_replay_count": "重聽的筆數",
    # node_10
    "automatable_count": "可自動化處理的筆數",
    "non_automatable_count": "不可自動化處理的筆數",
    # node_15
    "ivr_transfer_count": "轉接既有語音各項功能的筆數",
    "non_ivr_count": "非轉接既有語音功能的筆數",
    # exception_01
    "first_time_redirect_count": "第一次說明轉專員的筆數",
    "second_time_transfer_count": "第二次說明轉專員並轉接的筆數",
    "transfer_intent_detected_count": "偵測到轉專員意圖的筆數",
    # exception_02
    "negative_first_time_retry_count": "第一次理解錯誤(負向回覆)的筆數",
    "negative_second_time_transfer_count": "連續兩次理解錯誤(負向)並轉接的筆數",
    "other_first_time_replay_count": "第一次理解錯誤(重聽)的筆數",
    "other_second_time_transfer_count": "連續兩次理解錯誤(重聽)並轉接的筆數",
    # node_19
    "no_other_question_count": "客戶沒有其他問題的筆數",
    "has_card_question_count": "客戶有其他信用卡問題的筆數",
    "has_bank_question_count": "客戶有銀行問題的筆數",
    "note": "備註",
    # exception_04
    "transferred_to_human_count": "發生錯誤並轉接專員的筆數",
}

# 分布類 (dict) 欄位的子表標題
DISTRIBUTION_LABELS = {
    "intent_distribution": "意圖分布 (LLM 辨識)",
    "action_distribution": "後續動作分布",
    "ivr_function_distribution": "轉接語音功能分布 (voiceCommand)",
    "wrong_intent_distribution": "錯誤意圖分布",
    "timeout_node_distribution": "Time Out 發生的節點",
    "error_node_distribution": "ERROR 發生的節點",
    "error_type_distribution": "ERROR 代碼 / 類型",
}

# list 類欄位 (明細太長，總覽只顯示筆數 + 提示)
LIST_LABELS = {
    "customer_queries": "客戶口述問題 (詳見明細 Sheet)",
    "customer_ids": "涉及客戶 ID 清單",
    "error_details": "ERROR 詳情 (詳見明細 Sheet)",
}

# workflow 尚未完整實作的模組 → Sheet 頂端警示
INCOMPLETE_MODULES = {
    "node_18": "⚠ workflow 尚未完整實作 SMS 確認 / 電話選擇 / 發送結果節點，部分指標為 N/A。",
    "node_19": "⚠ workflow 尚未實作「是否有其他問題」節點，目前無明細資料。",
}

# 明細 Sheet 固定優先欄位順序
PRIORITY_COLUMNS = [
    "session_id",
    "customer_id",
    "node",
    "exception",
    "entered",
    "stay_duration_sec",
    "execution_time_ms",
]

# ── 樣式 ────────────────────────────────────────────────────────
TITLE_FONT = Font(name="Microsoft JhengHei", size=14, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FONT = Font(name="Microsoft JhengHei", size=12, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="4472C4")
SUBSECTION_FONT = Font(name="Microsoft JhengHei", size=11, bold=True)
SUBSECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(name="Microsoft JhengHei", size=10, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
NOTE_FONT = Font(name="Microsoft JhengHei", size=10, italic=True, color="808080")
CELL_FONT = Font(name="Microsoft JhengHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── 內部工具函式 ────────────────────────────────────────────────

def _flatten_cell(value):
    """把 dict / list 型值序列化為 Excel 可顯示字串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)
        except (TypeError, ValueError):
            return str(value)
    return value


def _safe_sheet_name(name: str) -> str:
    """Excel sheet 名稱不可含 : \\ / ? * [ ] 且長度 <= 31。"""
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:31]


def _auto_width(ws, max_width: int = 40):
    """簡單依內容長度調整欄寬。中文以 2 倍寬計算。"""
    if ws.max_row == 0 or ws.max_column == 0:
        return
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # MergedCell 沒有獨立 value (只有 top-left 有)，跳過即可
            v = cell.value
            if v is None:
                continue
            s = str(v)
            length = sum(2 if ord(c) > 127 else 1 for c in s)
            max_len = max(max_len, length)
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def _write_kv_row(ws, row: int, label: str, value, label_col: int = 1):
    """寫一列 key-value (2 欄)。"""
    k_cell = ws.cell(row=row, column=label_col, value=label)
    k_cell.font = CELL_FONT
    k_cell.alignment = LEFT_ALIGN
    k_cell.border = THIN_BORDER

    v = _flatten_cell(value)
    v_cell = ws.cell(row=row, column=label_col + 1, value=v)
    v_cell.font = CELL_FONT
    v_cell.alignment = LEFT_ALIGN
    v_cell.border = THIN_BORDER


def _write_section_header(ws, row: int, text: str, span: int = 2):
    """寫區塊標題 (大藍底白字)。"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = LEFT_ALIGN
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _write_subsection_header(ws, row: int, text: str, span: int = 2):
    """寫子標題 (淺藍底)。"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBSECTION_FONT
    cell.fill = SUBSECTION_FILL
    cell.alignment = LEFT_ALIGN
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


# ── 總覽 Sheet ──────────────────────────────────────────────────

def _write_summary_section(ws, start_row: int, mod_name: str,
                            aggregate_dict: dict) -> int:
    """
    把單一模組的 aggregate 結果以 key-value + 子表形式寫到總覽 Sheet。
    回傳下一個可用列號。
    """
    row = start_row
    _write_section_header(ws, row, f"═ {mod_name} ═", span=2)
    row += 1

    # 先區分 scalar / dict / list 欄位
    scalars = {}
    distributions = {}
    lists = {}

    for key, val in aggregate_dict.items():
        if key == "report_item":
            continue
        if isinstance(val, dict):
            distributions[key] = val
        elif isinstance(val, list):
            lists[key] = val
        else:
            scalars[key] = val

    # 寫 scalar (key-value)
    for key, val in scalars.items():
        label = FIELD_LABELS.get(key, key)
        display_val = val if val is not None else "N/A"
        _write_kv_row(ws, row, label, display_val)
        row += 1

    # 寫 list 類 (只顯示筆數 + 提示)
    for key, val in lists.items():
        label = LIST_LABELS.get(key, key)
        _write_kv_row(ws, row, label, f"共 {len(val)} 筆")
        row += 1

    # 寫 distribution 子表
    for key, dist in distributions.items():
        row += 1  # 空一列
        sub_title = DISTRIBUTION_LABELS.get(key, key)
        _write_subsection_header(ws, row, f"  ▸ {sub_title}", span=2)
        row += 1
        if not dist:
            _write_kv_row(ws, row, "(無資料)", "")
            row += 1
        else:
            # 按次數由多到少排序
            sorted_items = sorted(dist.items(), key=lambda kv: -kv[1] if isinstance(kv[1], (int, float)) else 0)
            # 表頭
            h1 = ws.cell(row=row, column=1, value="類別")
            h1.font = HEADER_FONT
            h1.fill = HEADER_FILL
            h1.alignment = CENTER_ALIGN
            h1.border = THIN_BORDER
            h2 = ws.cell(row=row, column=2, value="次數")
            h2.font = HEADER_FONT
            h2.fill = HEADER_FILL
            h2.alignment = CENTER_ALIGN
            h2.border = THIN_BORDER
            row += 1
            for cat, count in sorted_items:
                _write_kv_row(ws, row, str(cat) if cat is not None else "(空)", count)
                row += 1

    row += 1  # 區塊間空一列
    return row


def _write_overview_sheet(wb: Workbook, report: dict, report_modules: list):
    """寫總覽 Sheet: metadata + 每個模組的彙總指標。"""
    ws = wb.create_sheet(title="總覽", index=0)

    # 大標
    cell = ws.cell(row=1, column=1, value=report.get("report_title", "報表"))
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = LEFT_ALIGN
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.row_dimensions[1].height = 24

    row = 3
    data_range = report.get("data_range", {})
    meta_lines = [
        ("產出時間", report.get("generated_at", "")),
        ("資料時間範圍 (起)", data_range.get("earliest_execution", "")),
        ("資料時間範圍 (迄)", data_range.get("latest_execution", "")),
        ("總 execution 筆數", data_range.get("total_executions", 0)),
        ("  成功", data_range.get("success_count", 0)),
        ("  錯誤", data_range.get("error_count", 0)),
        ("  等待中", data_range.get("waiting_count", 0)),
    ]
    for label, val in meta_lines:
        _write_kv_row(ws, row, label, val)
        row += 1

    row += 1  # 空一列

    # 每個模組彙總
    summary = report.get("summary", {})
    for mod_info in report_modules:
        mod_id = mod_info["id"]
        mod_name = mod_info["name"]
        aggregate_dict = summary.get(mod_id, {}) or {}

        # node_18 / node_19 標註 workflow 未完整
        if mod_id in INCOMPLETE_MODULES:
            warn_cell = ws.cell(row=row, column=1, value=INCOMPLETE_MODULES[mod_id])
            warn_cell.font = NOTE_FONT
            warn_cell.alignment = LEFT_ALIGN
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1

        row = _write_summary_section(ws, row, mod_name, aggregate_dict)

    _auto_width(ws, max_width=50)


# ── 明細 Sheet ──────────────────────────────────────────────────

def _order_columns(records: list) -> list:
    """
    依優先欄位 + 字母序排出明細欄位順序。
    """
    all_keys = set()
    for r in records:
        all_keys.update(r.keys())

    ordered = []
    for col in PRIORITY_COLUMNS:
        if col in all_keys:
            ordered.append(col)
            all_keys.discard(col)
    ordered.extend(sorted(all_keys))
    return ordered


def _write_detail_sheet(wb: Workbook, mod_id: str, mod_name: str,
                         detail_records: list):
    """寫單一模組的明細 Sheet (一列一筆 execution)。"""
    # Sheet 名稱: 從 mod_name 去掉開頭編號點
    #   "7.說出問題"        → "07_說出問題"
    #   "例外1.客戶直接說明轉專員" → "例外01_轉專員" (簡縮)
    sheet_name_map = {
        "node_07": "07_說出問題",
        "node_08": "08_理解問題",
        "node_09": "09_與客戶確認問題",
        "node_10": "10_是否可自動化處理",
        "node_15": "15_是否為語音一站式",
        "node_18": "18_說答案後發送訊息SMS",
        "node_19": "19_是否有其他問題",
        "exception_01": "例外01_轉專員",
        "exception_02": "例外02_理解錯誤",
        "exception_03": "例外03_TimeOut",
        "exception_04": "例外04_系統ERROR",
    }
    sheet_name = _safe_sheet_name(sheet_name_map.get(mod_id, mod_id))
    ws = wb.create_sheet(title=sheet_name)

    # Row 1: 大標
    title_cell = ws.cell(row=1, column=1, value=mod_name)
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = LEFT_ALIGN
    ws.row_dimensions[1].height = 24

    current_row = 2

    # Row 2: 警示 (若 workflow 未完整)
    if mod_id in INCOMPLETE_MODULES:
        warn_cell = ws.cell(row=current_row, column=1,
                             value=INCOMPLETE_MODULES[mod_id])
        warn_cell.font = NOTE_FONT
        warn_cell.alignment = LEFT_ALIGN
        current_row += 1

    current_row += 1  # 空一列

    if not detail_records:
        msg_cell = ws.cell(row=current_row, column=1,
                            value="目前無明細資料。")
        msg_cell.font = NOTE_FONT
        msg_cell.alignment = LEFT_ALIGN
        ws.column_dimensions["A"].width = 60
        return

    # 決定欄位順序
    columns = _order_columns(detail_records)

    # 合併大標至所有欄位
    if len(columns) > 1:
        ws.merge_cells(start_row=1, start_column=1,
                        end_row=1, end_column=len(columns))

    # Header
    for i, col_key in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=i, value=col_key)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # 資料列
    for row_idx, record in enumerate(detail_records, start=current_row + 1):
        for col_idx, col_key in enumerate(columns, start=1):
            val = record.get(col_key)
            cell = ws.cell(row=row_idx, column=col_idx,
                            value=_flatten_cell(val))
            cell.font = CELL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER

    # 凍結首列 (欄位名稱)
    ws.freeze_panes = ws.cell(row=current_row + 1, column=1)

    _auto_width(ws, max_width=50)


# ── 主入口 ──────────────────────────────────────────────────────

def build_excel_report(report: dict, output_path: str) -> None:
    """
    將 pipeline 產出的 report dict 輸出為 .xlsx 檔案。

    Sheet 結構:
        - 總覽: report metadata + 每個模組的彙總指標
        - 每個 REPORT_MODULES 對應一張明細 Sheet
    """
    # 匯入此處以避免循環 import
    from pipeline import REPORT_MODULES

    wb = Workbook()
    # openpyxl 預設的 "Sheet" 移除
    default = wb.active
    wb.remove(default)

    # Sheet 1: 總覽
    _write_overview_sheet(wb, report, REPORT_MODULES)

    # Sheet 2..N: 每個模組的明細
    detail_records = report.get("detail_records", {})
    for mod_info in REPORT_MODULES:
        mod_id = mod_info["id"]
        mod_name = mod_info["name"]
        records = detail_records.get(mod_id, []) or []
        _write_detail_sheet(wb, mod_id, mod_name, records)

    wb.save(output_path)
